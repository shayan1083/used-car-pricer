from flask import Flask, request, jsonify
from playwright.async_api import async_playwright
import asyncio
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import nbformat
from nbconvert.preprocessors import ExecutePreprocessor
from sklearn.preprocessing import OneHotEncoder
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.svm import SVR
from sklearn.neighbors import KNeighborsRegressor


app = Flask(__name__)

URL = 'https://www.iseecars.com/'

async def check_webpage_loaded():
    try:
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(URL, wait_until="networkidle")
            await page.wait_for_load_state("load")
            page_source = await page.content()
            await page.close()
            await browser.close()
        return page_source
    except Exception as e:
        print(e)
        return 'error occured'

async def get_car_makes_models():
    try:
        print("get_car_makes_models")
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(URL, wait_until="networkidle")
            await page.wait_for_load_state("load")
            # extract car makes from dropdown
            makes = await page.eval_on_selector_all("#make option", "options => options.map(option => option.textContent.trim())")
            makes.pop(0)
            models = {}
            for make in makes:
                await page.locator('#make').select_option(make)
                models_for_make = await page.eval_on_selector_all("#model option", "options => options.map(option => option.textContent.trim())")
                models_for_make.pop(0)
                models[make] = models_for_make
            
            await page.close()
            await browser.close()
        print("done")
        return makes, models
    except Exception as e:
        print(e)
        return [], {}

async def get_car_trims(make,model,zip):
    try:
        print("get_car_trims")
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.goto(URL, wait_until="networkidle")
            await page.wait_for_load_state("load")
            await page.locator('#make').select_option(make)
            await page.locator('#model').select_option(model)
            await page.get_by_placeholder("ZIP Code").click()
            await page.get_by_placeholder("ZIP Code").press_sequentially(zip)
            await page.get_by_role("button", name="SEARCH").click()
            await page.wait_for_load_state("load") 
            await page.locator("#TrimContainer div").nth(1).click()
            trims = await page.eval_on_selector_all("#TrimContainer div ul li", "options => options.map(option => option.textContent.trim())")
            await page.close()
            await browser.close()
        print("done")
        return trims
    except Exception as e:
        print(e)
        return []
    
async def get_data(make,model,zip):
    try:
        print("get_data")
        async with async_playwright() as p:
            # open website
            browser = await p.chromium.launch(headless=True)
            context = await browser.new_context()
            page = await context.new_page()
            await page.goto(URL, wait_until="networkidle")
            await page.wait_for_load_state("load")
            # enter car information and search
            await page.locator('#make').select_option(make)
            await page.locator('#model').select_option(model)
            await page.get_by_placeholder("ZIP Code").click()
            await page.get_by_placeholder("ZIP Code").press_sequentially(zip) 
            await page.get_by_role("button", name="SEARCH").click()
            await page.wait_for_load_state("load") 
            await asyncio.sleep(5)
            # prepare excel sheet for input
            WB = load_workbook('./CarData.xlsx')
            WS = WB.active
            WS.delete_rows(1, WS.max_row)
            ymmm_string = 'year_make_model_mileage' + make
            p_string = 'price' + model

            WS.append([ymmm_string, p_string]) 
            # loop for going through the pages with car listings. will break loop when button is no longer found
            while True:
                # get the divs of the listings
                car_listings = await page.locator('article.article-search-result').all()
                # handle no cars
                if not car_listings:
                    print("No Cars Found")
                    break
                # go through each div
                for listing in car_listings:
                    # get the information
                    year_make_model_mileage = await listing.locator('h3 span.detailLink span').text_content(timeout=5000)
                    price = await listing.locator('div.col3 h4').text_content(timeout=5000)
                    # handle when price isnt in its usual location
                    if not price:
                        price = await listing.locator('ul.result-items').inner_text()       
                    # add the data to the excel sheet
                    WS.append([year_make_model_mileage, price])
                # search for next button
                next_button = page.get_by_text("Next »")
                # if no next button then break the loop since there are no more listings
                if await next_button.count() == 0:
                    break
                # click the next button and wait for results to load
                await page.get_by_text("Next »").click()
                await asyncio.sleep(5)
            # save the excel
            WB.save('./CarData.xlsx')
            # close tab and browser
            await page.close()
            await browser.close()
        print("done")
    except Exception as e:
        print(e)
        WB.save('./CarData.xlsx') # when you cant press the next button anymore, error is thrown but save the file
        print("done")

async def get_colors(context, listing, page, year_make_model_mileage):
    # check the listing-url attribute of the current listing to see if it redirects
    listing_url = await listing.get_attribute('listing-url')
    # if the listing_url contains redirect url, then prepare to open a new page
    if 'redirectUrl' in listing_url:
        print("open new page\n")
        async with context.expect_page() as new_page_info:
            await page.get_by_text(year_make_model_mileage).click() # Opens a new tab
        new_page = await new_page_info.value
        await new_page.wait_for_load_state()
        await asyncio.sleep(3)
        exterior_color_selector = 'div.cz-vdp-ctc-vs-col:has(span:has-text("Exterior Color")) >> span:nth-of-type(2)'
        interior_color_selector = 'div.cz-vdp-ctc-vs-col:has(span:has-text("Interior Color")) >> span:nth-of-type(2)'
        exterior_color = await page.locator('span:has-text("Exterior Color")').text_content()
        interior_color = await page.locator(interior_color_selector).text_content()
        await new_page.close()
    else:
        print("stay on same page\n")
    
    return exterior_color, interior_color

def parse_car_details(car_string, WORDS):
    parts = car_string.split(' - ')
    if len(parts) != 2:
        raise ValueError("Invalid format: Expected ' - ' separator")
    description = parts[0]
    mileage_str = parts[1]
    year = description.split(' ')[0]
 
    make = WORDS[0]
    model = WORDS[1]
    
    trim_start_index = description.find(model) + len(model) + 1
    trim = description[trim_start_index:]

    mileage = mileage_str.split(' ')[0].replace(',', '')

    return {
        'year': year,
        'make': make,
        'model': model,
        'trim': trim,
        'mileage': mileage
    }

def extract_capital_words(text):
    word=""
    for c in text:
        if c.isupper() or c.isnumeric():
            word = text[text.index(c):]
            break
    return word

def prepare():
    print("prepare")
    df = pd.read_excel('CarData.xlsx')
    col_names = df.columns
    extracted_words = [extract_capital_words(col) for col in col_names]
    print(extracted_words)
    col1 = col_names[0].replace(extracted_words[0],'')
    col2 = col_names[1].replace(extracted_words[1],'')
    df = df.rename(columns={col_names[0]:col1,col_names[1]:col2})
    df = df.dropna()
    data = pd.DataFrame(columns=['Year', 'Make', 'Model', 'Trim', 'Mileage', 'Price'])
    for index, row in df.iterrows():
        try:
            details = parse_car_details(row[0], extracted_words)
        except ValueError:
            continue
        price = str(row[1])
        if not price.__contains__('market price'):
            price = price.split('\n')[0]
            price = price.replace('$','')
            price = price.replace(' ', '')
            price = price.replace(',','')
            price = price.replace('DealerRating', '')
        else:
            price = price.replace('market price', '')
            price = price.replace('Dealer Rating', '')
            price = price.replace('Low Annual Miles', '')
            price = price.split('\n')[0]
            price = price.replace('$','')
            price = price.replace(' ', '')
            price = price.replace(',','')
        try:
            price = float(price)
        except ValueError:
            continue
        mileage = int(details['mileage'].replace(' mi', ''))
        data_trim = details['trim']
        data_trim = data_trim.strip()
        row = {
            'Year': int(details['year']),
            'Make': details['make'],
            'Model': details['model'],
            'Trim': data_trim,
            'Mileage': int(mileage),
            'Price': price
        }
        data = pd.concat([data, pd.DataFrame([row])])
    y = data['Price']
    X = data.drop(['Price', 'Make', 'Model'], axis=1)
    encoder = OneHotEncoder(sparse_output=False)
    encoded = pd.DataFrame(encoder.fit_transform(X[['Trim']]))
    encoded.index = X.index
    df_encoded = X.drop('Trim', axis=1)
    df_encoded = pd.concat([df_encoded, encoded], axis=1)
    scaler = StandardScaler()
    columns = ['Year', 'Mileage']
    df_encoded[columns] = scaler.fit_transform(df_encoded[columns])
    X = df_encoded
    X.columns = X.columns.astype(str)
    print("done")
    return X,y, scaler, encoder

async def predict_price(year,trim,mileage):
    print("predict_price")
    X, y, scaler, encoder = prepare()
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2,random_state=1)
    def train_evaluate_model(model):
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        mae = mean_absolute_error(y_test, y_pred)
        mse = mean_squared_error(y_test, y_pred)
        rmse = mse ** 0.5
        return mae, rmse
    lr = LinearRegression()
    mae_lr, rmse_lr = train_evaluate_model(lr)
    ridge = Ridge()
    mae_ridge, rmse_ridge = train_evaluate_model(ridge)
    lasso = Lasso()
    mae_lasso, rmse_lasso = train_evaluate_model(lasso)
    dt = DecisionTreeRegressor()
    mae_dt, rmse_dt = train_evaluate_model(dt)
    rf = RandomForestRegressor(n_estimators=150)
    mae_rf, rmse_rf = train_evaluate_model(rf)
    gb = GradientBoostingRegressor()
    mae_gb, rmse_gb = train_evaluate_model(gb)
    svr = SVR(kernel='rbf')
    mae_svr, rmse_svr = train_evaluate_model(svr)
    knn = KNeighborsRegressor(n_neighbors=10)
    mae_knn, rmse_knn = train_evaluate_model(knn)
    
    results = {
        'Model': ['Linear Regression', 'Ridge Regression', 'Lasso Regression', 'Decision Tree', 'Random Forest',
                'Gradient Boosting', 'SVR', 'KNN'],
        'MAE': [mae_lr, mae_ridge, mae_lasso, mae_dt, mae_rf, mae_gb, mae_svr, mae_knn],
        'RMSE': [rmse_lr, rmse_ridge, rmse_lasso, rmse_dt, rmse_rf, rmse_gb, rmse_svr, rmse_knn],
        'Model_Type': [lr, ridge, lasso, dt, rf, gb, svr, knn]
    }
    min_mae_index = results['MAE'].index(min(results['MAE']))
    best_model = results['Model_Type'][min_mae_index]
    model_name = results['Model'][min_mae_index]
    new_car = {
        'Year': year,
        'Mileage': mileage,
        'Trim': trim
    }
    new_car_df = pd.DataFrame([new_car])
    new_car_df.columns = new_car_df.columns.astype(str)
    new_car_df[['Year','Mileage']] = scaler.transform(new_car_df[['Year','Mileage']])
    encoded_trim = pd.DataFrame(encoder.transform(new_car_df[['Trim']]))
    new_car_prepared = pd.concat([new_car_df[['Year', 'Mileage']], encoded_trim], axis=1)
    new_car_prepared.columns = new_car_prepared.columns.astype(str)
    predicted_price = best_model.predict(new_car_prepared)
    print("done: ",model_name)
    return float(predicted_price)



@app.route('/', methods=['GET'])
def test():
    print("Home route accessed")
    return "Home route accessed", 200

@app.route('/source', methods=['GET'])
async def source():
    text = await check_webpage_loaded()
    return text, 200

# get makes and models for each make
@app.route('/makes-models', methods=['GET'])
async def car_makes_models():
    makes, models = await get_car_makes_models()
    return jsonify({'makes': makes, 'models': models})

# get trims for the chosen make and model
@app.route('/trims', methods=['GET'])
async def trims():
    make = request.args.get('make')
    model = request.args.get('model')
    zip = request.args.get('zip')
    if not make or not model or not zip:
        return jsonify({'error': 'Make, Model, And Zip Code Are Required'}), 400
    trims = await get_car_trims(make,model,zip)
    return jsonify({'trims': trims})

# get car data based on chosen info and put into excel
@app.route('/data', methods=['GET'])
async def data():
    make = request.args.get('make')
    model = request.args.get('model')
    zip = request.args.get('zip')
    #trim = request.args.get('trim')
    await get_data(make,model,zip)
    return 'Data in excel file', 200

# perform analysis on data from excel file and predict a price
@app.route('/predict', methods=['GET'])
async def predict():
    year = request.args.get('year')
    trim = request.args.get('trim')
    miles = request.args.get('miles')
    # scrape data and perform machine learning'
    predicted_value = await predict_price(year,trim,miles)
    return jsonify({'predicted_value': predicted_value})

if __name__ == '__main__':
    app.run(debug=True)